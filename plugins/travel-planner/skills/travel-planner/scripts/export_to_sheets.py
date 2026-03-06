#!/usr/bin/env python3
"""
Export Italy trip plan markdown to a formatted Excel file (.xlsx) that can be
uploaded to Google Sheets.  Uses openpyxl for rich formatting — no Google API
credentials required.

Usage:
    python3 export_to_sheets.py [input.md] [output.xlsx]

Defaults:
    input  = italy-summer-2026-plan-v2.md (same directory as script)
    output = italy-summer-2026-plan.xlsx
"""

import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing dependency. Install with:\n  pip3 install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Style palette
# ---------------------------------------------------------------------------
COLORS = {
    "header_bg": "1B4F72",      # dark blue
    "header_fg": "FFFFFF",
    "section_bg": "2E86C1",     # medium blue
    "section_fg": "FFFFFF",
    "subsection_bg": "AED6F1",  # light blue
    "subsection_fg": "1B4F72",
    "day_bg": "F39C12",         # warm orange
    "day_fg": "FFFFFF",
    "planb_bg": "FEF9E7",       # pale yellow
    "planb_fg": "7D6608",
    "table_header_bg": "D5F5E3", # light green
    "table_header_fg": "1B4F72",
    "alt_row": "F2F4F4",        # very light grey
    "white": "FFFFFF",
    "border": "BDC3C7",
}

thin_border = Border(
    left=Side(style="thin", color=COLORS["border"]),
    right=Side(style="thin", color=COLORS["border"]),
    top=Side(style="thin", color=COLORS["border"]),
    bottom=Side(style="thin", color=COLORS["border"]),
)


def make_font(bold=False, size=11, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)


def make_fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


# ---------------------------------------------------------------------------
# Markdown parser — extracts structured sections from the plan
# ---------------------------------------------------------------------------
def parse_markdown(text: str) -> list[dict]:
    """Return a list of content blocks: headings, tables, bullets, quotes, text."""
    blocks: list[dict] = []
    lines = text.split("\n")
    i = 0

    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        # Skip empty lines and horizontal rules
        if not stripped or stripped == "---":
            i += 1
            continue

        # Headings
        if stripped.startswith("#"):
            level = len(stripped) - len(stripped.lstrip("#"))
            title = stripped.lstrip("#").strip()
            blocks.append({"type": "heading", "level": level, "text": title})
            i += 1
            continue

        # Blockquotes
        if stripped.startswith(">"):
            quote_lines = []
            while i < len(lines) and lines[i].strip().startswith(">"):
                quote_lines.append(lines[i].strip().lstrip(">").strip())
                i += 1
            blocks.append({"type": "quote", "text": "\n".join(quote_lines)})
            continue

        # Tables
        if "|" in stripped and stripped.startswith("|"):
            table_lines = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                row = lines[i].strip()
                # Skip separator rows (|---|---|)
                if not re.match(r"^\|[\s\-:|]+\|$", row):
                    cells = [c.strip() for c in row.split("|")[1:-1]]
                    table_lines.append(cells)
                i += 1
            if table_lines:
                blocks.append({"type": "table", "rows": table_lines})
            continue

        # Bullet points
        if stripped.startswith("- "):
            bullets = []
            while i < len(lines) and (lines[i].strip().startswith("- ") or lines[i].strip().startswith("  - ")):
                bullets.append(lines[i].strip().lstrip("- ").strip())
                i += 1
            blocks.append({"type": "bullets", "items": bullets})
            continue

        # Checkbox items
        if stripped.startswith("- [ ]") or stripped.startswith("- [x]"):
            checks = []
            while i < len(lines) and (lines[i].strip().startswith("- [ ]") or lines[i].strip().startswith("- [x]")):
                item = lines[i].strip()
                done = item.startswith("- [x]")
                text = item[5:].strip()
                checks.append({"done": done, "text": text})
                i += 1
            blocks.append({"type": "checklist", "items": checks})
            continue

        # Bold standalone lines (like **Recommended: ...** )
        if stripped.startswith("**") and stripped.endswith("**"):
            blocks.append({"type": "bold_line", "text": stripped.strip("*").strip()})
            i += 1
            continue

        # Regular text
        blocks.append({"type": "text", "text": stripped})
        i += 1

    return blocks


def clean_md(text: str) -> str:
    """Strip markdown formatting for cell display."""
    # Remove links but keep text: [text](url) → text
    text = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", text)
    # Remove bold/italic markers
    text = text.replace("**", "").replace("__", "")
    text = re.sub(r"(?<!\*)\*(?!\*)", "", text)
    # Remove inline code
    text = re.sub(r"`([^`]+)`", r"\1", text)
    return text.strip()


def extract_url(text: str) -> str | None:
    """Extract first URL from markdown link."""
    m = re.search(r"\[([^\]]+)\]\(([^)]+)\)", text)
    return m.group(2) if m else None


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------
def write_excel(blocks: list[dict], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Italy Trip Plan"
    ws.sheet_properties.tabColor = COLORS["header_bg"]

    # Column widths
    ws.column_dimensions["A"].width = 85
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 25

    row = 1

    def write_cell(r, c, value, font=None, fill=None, alignment=None, border=None):
        cell = ws.cell(row=r, column=c, value=clean_md(str(value)))
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        else:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if border:
            cell.border = border
        # Add hyperlink if markdown link detected
        url = extract_url(str(value))
        if url:
            cell.hyperlink = url
            cell.font = Font(
                name="Calibri",
                bold=font.bold if font else False,
                size=font.size if font else 11,
                color="2E86C1",
                underline="single",
            )

    in_plan_b = False
    in_day = False

    for block in blocks:
        btype = block["type"]

        if btype == "heading":
            level = block["level"]
            text = block["text"]

            in_plan_b = "Plan B" in text
            is_day = text.startswith("Day ")
            if is_day:
                in_day = True

            if level == 1:
                # Title row — merge across columns, big font
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                write_cell(row, 1, text,
                           font=make_font(bold=True, size=18, color=COLORS["header_fg"]),
                           fill=make_fill(COLORS["header_bg"]),
                           alignment=Alignment(horizontal="center", vertical="center"))
                ws.row_dimensions[row].height = 40
                row += 1

            elif level == 2:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                write_cell(row, 1, text,
                           font=make_font(bold=True, size=14, color=COLORS["section_fg"]),
                           fill=make_fill(COLORS["section_bg"]),
                           alignment=Alignment(vertical="center"))
                ws.row_dimensions[row].height = 30
                row += 1

            elif level == 3:
                if is_day:
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                    write_cell(row, 1, text,
                               font=make_font(bold=True, size=13, color=COLORS["day_fg"]),
                               fill=make_fill(COLORS["day_bg"]),
                               alignment=Alignment(vertical="center"))
                    ws.row_dimensions[row].height = 28
                else:
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                    write_cell(row, 1, text,
                               font=make_font(bold=True, size=12, color=COLORS["subsection_fg"]),
                               fill=make_fill(COLORS["subsection_bg"]))
                    ws.row_dimensions[row].height = 25
                row += 1

            elif level >= 4:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                write_cell(row, 1, text,
                           font=make_font(bold=True, size=11, color=COLORS["subsection_fg"]))
                row += 1

        elif btype == "table":
            rows_data = block["rows"]
            if not rows_data:
                continue
            num_cols = len(rows_data[0])

            # Header row
            for ci, cell_val in enumerate(rows_data[0], 1):
                write_cell(row, ci, cell_val,
                           font=make_font(bold=True, size=10, color=COLORS["table_header_fg"]),
                           fill=make_fill(COLORS["table_header_bg"]),
                           border=thin_border)
            row += 1

            # Data rows
            for ri, data_row in enumerate(rows_data[1:]):
                fill = make_fill(COLORS["alt_row"]) if ri % 2 == 1 else make_fill(COLORS["white"])
                for ci, cell_val in enumerate(data_row, 1):
                    write_cell(row, ci, cell_val,
                               font=make_font(size=10),
                               fill=fill,
                               border=thin_border)
                row += 1
            row += 1  # blank row after table

        elif btype == "bullets":
            for item in block["items"]:
                fill = make_fill(COLORS["planb_bg"]) if in_plan_b else None
                font_color = COLORS["planb_fg"] if in_plan_b else "000000"
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                prefix = "  " if in_plan_b else ""
                write_cell(row, 1, f"{prefix}• {item}",
                           font=make_font(size=10, color=font_color, italic=in_plan_b),
                           fill=fill)
                row += 1

        elif btype == "checklist":
            for item in block["items"]:
                marker = "☑" if item["done"] else "☐"
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                write_cell(row, 1, f"  {marker}  {item['text']}",
                           font=make_font(size=10))
                row += 1

        elif btype == "quote":
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            write_cell(row, 1, block["text"],
                       font=make_font(size=10, italic=True, color="5D6D7E"),
                       fill=make_fill("EBF5FB"))
            row += 1

        elif btype == "bold_line":
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            write_cell(row, 1, block["text"],
                       font=make_font(bold=True, size=11))
            row += 1

        elif btype == "text":
            text = block["text"]
            if not text:
                continue
            fill = make_fill(COLORS["planb_bg"]) if in_plan_b else None
            font_color = COLORS["planb_fg"] if in_plan_b else "000000"
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            write_cell(row, 1, text,
                       font=make_font(size=10, color=font_color),
                       fill=fill)
            row += 1

    # Freeze top row
    ws.freeze_panes = "A2"

    # Print settings
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    wb.save(output_path)
    print(f"Exported {row - 1} rows → {output_path}")
    print(f"Upload to Google Sheets: drive.google.com → New → File upload → select {Path(output_path).name}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    script_dir = Path(__file__).parent
    input_path = sys.argv[1] if len(sys.argv) > 1 else str(script_dir / "italy-summer-2026-plan-v2.md")
    output_path = sys.argv[2] if len(sys.argv) > 2 else str(script_dir / "italy-summer-2026-plan.xlsx")

    md_text = Path(input_path).read_text(encoding="utf-8")
    blocks = parse_markdown(md_text)
    write_excel(blocks, output_path)


if __name__ == "__main__":
    main()
